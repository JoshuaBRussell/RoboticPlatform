#Script to aggregate the data summary files produced by the Matlab analysis code
#into a single sheet. 

import pandas as pd
from openpyxl import load_workbook


OUTPUT_FILE_NAME = './results/DIR_Summary.xlsx'
SUBJ_NAMES = ["Carl", "Vu", "Emily", "Ashley", "Carly", "Lily", "Matt", "James", "Kwanghee"]

RESULTS_DIR = r'./results/'

FILE_NAME_SUFFIX = "_datasummary.xlsx"



for i in range(len(SUBJ_NAMES)):
    curr_file =  RESULTS_DIR + SUBJ_NAMES[i] + '/' + SUBJ_NAMES[i]+FILE_NAME_SUFFIX
    print("Opening: ", curr_file)

    spread_sheet_obj = pd.ExcelFile(curr_file)
    
    book = load_workbook(OUTPUT_FILE_NAME)
    writer = pd.ExcelWriter(OUTPUT_FILE_NAME, mode='w', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    spread_sheet_obj.parse().to_excel(writer, index=False,sheet_name = 'Sheet1', startrow=(writer.sheets['Sheet1'].max_row + 1))

    writer.save()